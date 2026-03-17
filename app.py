#!/usr/bin/env python3
import argparse
import csv
import io
import json
import re
import socket
import threading
import zipfile
from dataclasses import dataclass, field
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).parent
STATIC_DIR = ROOT_DIR / "static"
MAX_UPLOAD_BYTES = 120 * 1024 * 1024
REGISTER_PATTERN = re.compile(r"^(DM|R)\d+$")
EVENT_RANGE_PATTERN = re.compile(r"(\d{4,6})\s*[-–]\s*(\d{4,6})")

OUTPUT_SIGNAL_NAMES = {
    0: "Output Signal #1 (00608 pins 3&4)",
    1: "Output Signal #2 (00609 pins 5&6)",
    2: "Output Signal #3 (00610 pins 7&8)",
    3: "Output Signal #4 (00611 pins 9&10)",
    4: "Output Signal #5 (00612 pins 11&12)",
    5: "Output Signal #6 (00613 pins 13&14)",
    6: "Output Signal #7 (00614 pins 15&16)",
    7: "Output Signal #8 (00615 pins 17&18)",
    8: "Output Signal #9 (00700 pins 19&20)",
    9: "Output Signal #10 (00701 pins 21&22)",
    10: "Output Signal #11 (00702 pins 23&24)",
}


def decode_escapes(value: str) -> str:
    return (
        value.replace("\\r", "\r")
        .replace("\\n", "\n")
        .replace("\\t", "\t")
        .replace("\\0", "\0")
    )


def parse_first_int(response: str) -> Optional[int]:
    match = re.search(r"[-+]?\d+", response)
    if not match:
        return None
    return int(match.group(0))


@dataclass
class TcpConnectionConfig:
    host: str
    port: int
    timeout_seconds: float


@dataclass
class ProtocolConfig:
    read_template: str
    write_template: str
    terminator: str
    encoding: str


class SimulatorStore:
    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._registers: Dict[str, int] = {
            "DM0": 0,
            "DM1": 100,
            "DM2": 200,
            "R0": 0,
            "R1": 1,
            "R2": 0,
        }

    def read(self, register: str) -> int:
        register = register.strip().upper()
        with self._lock:
            return self._registers.get(register, 0)

    def write(self, register: str, value: int) -> None:
        register = register.strip().upper()
        with self._lock:
            self._registers[register] = value

    def snapshot(self) -> Dict[str, int]:
        with self._lock:
            return dict(sorted(self._registers.items()))

    def load_registers(self, registers: Dict[str, int]) -> int:
        sanitized = {}
        for register, value in registers.items():
            key = register.strip().upper()
            if not REGISTER_PATTERN.match(key):
                continue
            sanitized[key] = int(value)

        with self._lock:
            self._registers.update(sanitized)
        return len(sanitized)


@dataclass
class EventDefinition:
    sequence: str
    label: str
    start: int
    end: int
    row_number: int
    event_type_code: Optional[int]
    alarm_items: List[str] = field(default_factory=list)


class EventDefinitionStore:
    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._definitions: List[EventDefinition] = [
            EventDefinition(
                sequence="Capture Canister - Removal",
                label="Event Type - Step 1",
                start=23480,
                end=23489,
                row_number=2349,
                event_type_code=135,
                alarm_items=[],
            )
        ]

    def get_all(self) -> List[EventDefinition]:
        with self._lock:
            return list(self._definitions)

    def replace(self, new_definitions: List[EventDefinition]) -> None:
        with self._lock:
            self._definitions = list(new_definitions)


class PlcTcpClient:
    def __init__(self, conn: TcpConnectionConfig, protocol: ProtocolConfig) -> None:
        self.conn = conn
        self.protocol = protocol

    def read(self, register: str) -> Dict[str, Any]:
        command = self.protocol.read_template.format(register=register)
        response = self._send(command)
        return {
            "command": command,
            "rawResponse": response,
            "parsedValue": parse_first_int(response),
        }

    def write(self, register: str, value: int) -> Dict[str, Any]:
        command = self.protocol.write_template.format(register=register, value=value)
        response = self._send(command)
        return {
            "command": command,
            "rawResponse": response,
        }

    def _send(self, command: str) -> str:
        payload = command + self.protocol.terminator
        encoded = payload.encode(self.protocol.encoding)

        with socket.create_connection(
            (self.conn.host, self.conn.port), timeout=self.conn.timeout_seconds
        ) as plc_socket:
            plc_socket.sendall(encoded)
            plc_socket.settimeout(self.conn.timeout_seconds)
            chunks = []
            while True:
                try:
                    data = plc_socket.recv(4096)
                except socket.timeout:
                    break
                if not data:
                    break
                chunks.append(data)
                if len(data) < 4096:
                    break

        if not chunks:
            return ""
        return b"".join(chunks).decode(self.protocol.encoding, errors="replace")


SIMULATOR = SimulatorStore()
EVENT_DEFINITIONS = EventDefinitionStore()


def parse_numeric_value(raw_value: str) -> Optional[int]:
    value = raw_value.strip()
    if not value or value == "-":
        return None

    if value.startswith("$"):
        try:
            return int(value[1:], 16)
        except ValueError:
            return None

    try:
        return int(value)
    except ValueError:
        try:
            return int(float(value))
        except ValueError:
            return None


def decode_csv_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not decode CSV content")


def parse_plc_device_csv(raw_csv: bytes) -> Tuple[Dict[str, int], Dict[str, int]]:
    content = decode_csv_bytes(raw_csv)
    reader = csv.reader(io.StringIO(content))

    registers: Dict[str, int] = {}
    matched_rows = 0
    loaded_rows = 0
    skipped_value_rows = 0

    for row in reader:
        if len(row) < 3:
            continue

        device = row[1].strip().upper()
        if not REGISTER_PATTERN.match(device):
            continue

        matched_rows += 1
        parsed_value = parse_numeric_value(row[2])
        if parsed_value is None:
            skipped_value_rows += 1
            continue

        registers[device] = parsed_value
        loaded_rows += 1

    return registers, {
        "matchedRows": matched_rows,
        "loadedRows": loaded_rows,
        "skippedRows": skipped_value_rows,
    }


def parse_uploaded_register_file(
    filename: str, payload: bytes
) -> Tuple[Dict[str, int], Dict[str, Any]]:
    suffix = Path(filename).suffix.lower()
    source_file = filename
    csv_bytes = payload

    if suffix == ".zip" or payload.startswith(b"PK"):
        with zipfile.ZipFile(io.BytesIO(payload), mode="r") as zip_file:
            names = zip_file.namelist()
            preferred = [name for name in names if name.lower().endswith("plcdevicevalue.csv")]
            if preferred:
                source_file = preferred[0]
            else:
                candidates = [name for name in names if name.lower().endswith(".csv")]
                if not candidates:
                    raise ValueError("ZIP does not contain a CSV file")
                source_file = candidates[0]
            csv_bytes = zip_file.read(source_file)
    elif suffix != ".csv":
        raise ValueError("Only .zip and .csv files are supported")

    registers, stats = parse_plc_device_csv(csv_bytes)
    metadata: Dict[str, Any] = {
        "sourceFile": source_file,
        "fileName": filename,
        **stats,
    }
    return registers, metadata


def is_register_address_candidate(value: int) -> bool:
    return 1000 <= value <= 999999


def parse_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    return None


def parse_event_definitions_from_rows(rows: List[List[Any]]) -> List[EventDefinition]:
    if not rows:
        return []

    grouped: Dict[Tuple[str, int], Dict[str, Any]] = {}
    for values in rows[1:]:
        if len(values) <= 10:
            continue

        row_number = parse_int(values[0] if len(values) > 0 else None)
        column_letter = str(values[1]).strip().upper() if len(values) > 1 and values[1] is not None else ""
        register = parse_int(values[2] if len(values) > 2 else None)
        row_label = str(values[4]).strip() if len(values) > 4 and values[4] is not None else ""
        decimal_value = parse_int(values[6] if len(values) > 6 else None)
        event_type_code = parse_int(values[7] if len(values) > 7 else None)
        sequence = str(values[10]).strip() if len(values) > 10 and values[10] is not None else ""

        if not sequence:
            continue
        if register is None or not is_register_address_candidate(register):
            continue
        if column_letter not in {"C", "D", "E", "F", "G", "H", "I", "J", "K", "L"}:
            continue
        if row_number is None:
            # Rows in this workbook always include row id; fallback keeps group stable.
            row_number = register // 10

        key = (sequence, row_number)
        item = grouped.setdefault(
            key,
            {
                "registers": [],
                "labels": [],
                "eventTypeCodes": [],
                "firstDecimalValue": None,
                "alarmItems": set(),
            },
        )
        item["registers"].append(register)
        if row_label:
            item["labels"].append(row_label)
        if event_type_code is not None:
            item["eventTypeCodes"].append(event_type_code)
        if column_letter == "C" and decimal_value is not None:
            item["firstDecimalValue"] = decimal_value
        alarm_code = parse_int(values[15] if len(values) > 15 else None)
        alarm_text = (
            str(values[16]).strip()
            if len(values) > 16 and values[16] is not None
            else ""
        )
        if alarm_code is not None or alarm_text:
            if alarm_text:
                item["alarmItems"].add(f"{alarm_code if alarm_code is not None else '?'}: {alarm_text}")
            elif alarm_code is not None:
                item["alarmItems"].add(str(alarm_code))

    parsed: List[EventDefinition] = []
    for (sequence, row_number), item in grouped.items():
        registers = sorted(set(item["registers"]))
        if len(registers) < 2:
            continue

        start = min(registers)
        end = max(registers)
        if end - start > 25:
            continue

        label = item["labels"][0] if item["labels"] else f"Row {row_number}"
        if item["eventTypeCodes"]:
            event_code = item["eventTypeCodes"][0]
        else:
            event_code = item["firstDecimalValue"]

        parsed.append(
            EventDefinition(
                sequence=sequence,
                label=label,
                start=start,
                end=end,
                row_number=row_number,
                event_type_code=event_code,
                alarm_items=sorted(item["alarmItems"]),
            )
        )

    parsed.sort(key=lambda item: (item.sequence.lower(), item.start, item.row_number))
    return parsed


def parse_event_definitions_from_csv(raw_bytes: bytes) -> List[EventDefinition]:
    content = decode_csv_bytes(raw_bytes)
    reader = csv.reader(io.StringIO(content))
    rows = [list(row) for row in reader]
    parsed = parse_event_definitions_from_rows(rows)
    if parsed:
        return parsed

    # Fallback for simplified range-only CSV
    range_based: List[EventDefinition] = []
    for row in rows:
        cells = [str(cell).strip() for cell in row if str(cell).strip()]
        if not cells:
            continue
        joined = " | ".join(cells)
        match = EVENT_RANGE_PATTERN.search(joined)
        if not match:
            continue
        start = int(match.group(1))
        end = int(match.group(2))
        if start > end:
            start, end = end, start
        name = next((cell for cell in cells if re.search(r"[A-Za-z]", cell)), "Unknown")
        range_based.append(
            EventDefinition(
                sequence=name,
                label="Range",
                start=start,
                end=end,
                row_number=start // 10,
                event_type_code=None,
                alarm_items=[],
            )
        )
    range_based.sort(key=lambda item: (item.sequence.lower(), item.start))
    return range_based


def parse_event_definitions_from_xlsx(raw_bytes: bytes) -> List[EventDefinition]:
    workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
    parsed: List[EventDefinition] = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheet_parsed = parse_event_definitions_from_rows(rows)
        parsed.extend(sheet_parsed)
    parsed.sort(key=lambda item: (item.sequence.lower(), item.start, item.row_number))
    return parsed


def parse_uploaded_event_list_file(filename: str, payload: bytes) -> List[EventDefinition]:
    suffix = Path(filename).suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return parse_event_definitions_from_xlsx(payload)
    if suffix == ".csv":
        return parse_event_definitions_from_csv(payload)
    raise ValueError("Only .xlsx, .xlsm, and .csv files are supported for event list")


def load_default_event_definitions() -> None:
    candidates = [
        ROOT_DIR / "Registers_RevM_EventTypeList_WithAlarmText.xlsx",
        ROOT_DIR / "Registers_RevM_EventTypeList_WithAlarmText.csv",
    ]
    for path in candidates:
        if not path.exists():
            continue
        try:
            definitions = parse_uploaded_event_list_file(path.name, path.read_bytes())
            if definitions:
                EVENT_DEFINITIONS.replace(definitions)
                print(
                    f"Loaded default event definitions from {path.name} "
                    f"({len(definitions)} blocks)"
                )
                return
        except Exception as exc:
            print(f"Warning: unable to load default event definitions from {path.name}: {exc}")


def decode_active_bits(words: List[int]) -> List[int]:
    active = []
    for word_index, word in enumerate(words):
        value = int(word) & 0xFFFF
        for bit in range(16):
            if value & (1 << bit):
                active.append(word_index * 16 + bit)
    return active


def decode_event_definition(definition: EventDefinition, registers: Dict[str, int]) -> Dict[str, Any]:
    words = [int(registers.get(f"DM{addr}", 0)) for addr in range(definition.start, definition.end + 1)]
    event_type_value = words[0] if words else 0
    output_words = words[1:5] if len(words) > 1 else []
    input_words = words[5:9] if len(words) > 5 else []
    control_words = words[9:] if len(words) > 9 else []

    output_bits = decode_active_bits(output_words)
    input_bits = decode_active_bits(input_words)
    named_outputs = [
        OUTPUT_SIGNAL_NAMES[bit] for bit in output_bits if bit in OUTPUT_SIGNAL_NAMES
    ]
    outputs_on = named_outputs + [
        f"Output Bit {bit}" for bit in output_bits if bit not in OUTPUT_SIGNAL_NAMES
    ]
    inputs_on = [f"Input Bit {bit}" for bit in input_bits]

    return {
        "sequence": definition.sequence,
        "label": definition.label,
        "start": definition.start,
        "end": definition.end,
        "rowNumber": definition.row_number,
        "eventTypeCode": definition.event_type_code,
        "wordCount": len(words),
        "eventTypeValue": event_type_value,
        "rawWords": words,
        "outputWords": output_words,
        "inputWords": input_words,
        "controlWords": control_words,
        "activeOutputBits": output_bits,
        "activeInputBits": input_bits,
        "namedActiveOutputs": named_outputs,
        "outputsOn": outputs_on,
        "inputsOn": inputs_on,
        "alarmItems": definition.alarm_items,
    }


class RequestHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(ROOT_DIR), **kwargs)

    def do_GET(self) -> None:
        if self.path == "/":
            self.path = "/static/index.html"
            return super().do_GET()

        if self.path == "/api/simulator/registers":
            return self._send_json(HTTPStatus.OK, {"registers": SIMULATOR.snapshot()})
        if self.path == "/api/event-definitions":
            return self._handle_get_event_definitions()

        return super().do_GET()

    def do_POST(self) -> None:
        if self.path == "/api/read":
            return self._handle_read()
        if self.path == "/api/write":
            return self._handle_write()
        if self.path == "/api/upload-register-file":
            return self._handle_upload_register_file()
        if self.path == "/api/upload-event-list":
            return self._handle_upload_event_list()
        return self._send_json(HTTPStatus.NOT_FOUND, {"error": "Unknown endpoint"})

    def _handle_read(self) -> None:
        payload = self._read_json_body()
        if payload is None:
            return

        register = str(payload.get("register", "")).strip().upper()
        if not register:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "register is required"}
            )

        mode = str(payload.get("mode", "simulator")).strip().lower()
        if mode == "simulator":
            value = SIMULATOR.read(register)
            return self._send_json(
                HTTPStatus.OK,
                {"mode": "simulator", "register": register, "value": value},
            )

        if mode == "tcp":
            try:
                client = self._build_tcp_client(payload)
                result = client.read(register)
            except Exception as exc:  # pragma: no cover - endpoint safety
                return self._send_json(
                    HTTPStatus.BAD_REQUEST,
                    {"error": f"TCP read failed: {exc}"},
                )

            return self._send_json(
                HTTPStatus.OK,
                {
                    "mode": "tcp",
                    "register": register,
                    "value": result["parsedValue"],
                    "rawResponse": result["rawResponse"],
                    "command": result["command"],
                },
            )

        return self._send_json(
            HTTPStatus.BAD_REQUEST, {"error": f"Unsupported mode: {mode}"}
        )

    def _handle_write(self) -> None:
        payload = self._read_json_body()
        if payload is None:
            return

        register = str(payload.get("register", "")).strip().upper()
        if not register:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "register is required"}
            )

        try:
            value = int(payload.get("value"))
        except (TypeError, ValueError):
            return self._send_json(
                HTTPStatus.BAD_REQUEST,
                {"error": "value is required and must be an integer"},
            )

        mode = str(payload.get("mode", "simulator")).strip().lower()
        if mode == "simulator":
            SIMULATOR.write(register, value)
            return self._send_json(
                HTTPStatus.OK,
                {"mode": "simulator", "register": register, "value": value},
            )

        if mode == "tcp":
            try:
                client = self._build_tcp_client(payload)
                result = client.write(register, value)
            except Exception as exc:  # pragma: no cover - endpoint safety
                return self._send_json(
                    HTTPStatus.BAD_REQUEST,
                    {"error": f"TCP write failed: {exc}"},
                )

            return self._send_json(
                HTTPStatus.OK,
                {
                    "mode": "tcp",
                    "register": register,
                    "value": value,
                    "rawResponse": result["rawResponse"],
                    "command": result["command"],
                },
            )

        return self._send_json(
            HTTPStatus.BAD_REQUEST, {"error": f"Unsupported mode: {mode}"}
        )

    def _build_tcp_client(self, payload: Dict[str, Any]) -> PlcTcpClient:
        connection = payload.get("connection") or {}
        protocol = payload.get("protocol") or {}

        host = str(connection.get("host", "")).strip()
        if not host:
            raise ValueError("connection.host is required for TCP mode")

        port = int(connection.get("port", 8501))
        timeout_seconds = float(connection.get("timeoutSeconds", 1.5))
        if timeout_seconds <= 0:
            raise ValueError("connection.timeoutSeconds must be positive")

        read_template = str(protocol.get("readTemplate", "RD {register}"))
        write_template = str(protocol.get("writeTemplate", "WR {register} {value}"))
        terminator = decode_escapes(str(protocol.get("terminator", "\\r\\n")))
        encoding = str(protocol.get("encoding", "ascii"))

        client = PlcTcpClient(
            conn=TcpConnectionConfig(
                host=host,
                port=port,
                timeout_seconds=timeout_seconds,
            ),
            protocol=ProtocolConfig(
                read_template=read_template,
                write_template=write_template,
                terminator=terminator,
                encoding=encoding,
            ),
        )
        return client

    def _handle_upload_register_file(self) -> None:
        content_len_header = self.headers.get("Content-Length")
        if not content_len_header:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "Missing Content-Length"}
            )

        try:
            content_length = int(content_len_header)
        except ValueError:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "Invalid Content-Length"}
            )

        if content_length <= 0:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "Uploaded file is empty"}
            )

        if content_length > MAX_UPLOAD_BYTES:
            return self._send_json(
                HTTPStatus.BAD_REQUEST,
                {
                    "error": (
                        f"File is too large ({content_length} bytes). "
                        f"Max supported size is {MAX_UPLOAD_BYTES} bytes."
                    )
                },
            )

        raw_payload = self.rfile.read(content_length)
        filename = Path(self.headers.get("X-Filename", "upload.bin")).name

        try:
            registers, metadata = parse_uploaded_register_file(filename, raw_payload)
            loaded_count = SIMULATOR.load_registers(registers)
        except Exception as exc:
            return self._send_json(
                HTTPStatus.BAD_REQUEST,
                {"error": f"Unable to import register file: {exc}"},
            )

        return self._send_json(
            HTTPStatus.OK,
            {
                "message": "Simulator registers imported successfully",
                "loadedRegisters": loaded_count,
                "sourceFile": metadata["sourceFile"],
                "matchedRows": metadata["matchedRows"],
                "loadedRows": metadata["loadedRows"],
                "skippedRows": metadata["skippedRows"],
            },
        )

    def _handle_upload_event_list(self) -> None:
        content_len_header = self.headers.get("Content-Length")
        if not content_len_header:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "Missing Content-Length"}
            )

        try:
            content_length = int(content_len_header)
        except ValueError:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "Invalid Content-Length"}
            )

        if content_length <= 0:
            return self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": "Uploaded file is empty"}
            )
        if content_length > MAX_UPLOAD_BYTES:
            return self._send_json(
                HTTPStatus.BAD_REQUEST,
                {
                    "error": (
                        f"File is too large ({content_length} bytes). "
                        f"Max supported size is {MAX_UPLOAD_BYTES} bytes."
                    )
                },
            )

        raw_payload = self.rfile.read(content_length)
        filename = Path(self.headers.get("X-Filename", "event-list.bin")).name

        try:
            definitions = parse_uploaded_event_list_file(filename, raw_payload)
            if not definitions:
                raise ValueError("No event/register ranges found in file")
            EVENT_DEFINITIONS.replace(definitions)
        except Exception as exc:
            return self._send_json(
                HTTPStatus.BAD_REQUEST,
                {"error": f"Unable to import event list file: {exc}"},
            )

        return self._send_json(
            HTTPStatus.OK,
            {
                "message": "Event list imported successfully",
                "loadedBlocks": len(definitions),
                "loadedSequences": len({item.sequence for item in definitions}),
                "firstBlock": {
                    "sequence": definitions[0].sequence,
                    "label": definitions[0].label,
                    "start": definitions[0].start,
                    "end": definitions[0].end,
                },
            },
        )

    def _handle_get_event_definitions(self) -> None:
        definitions = EVENT_DEFINITIONS.get_all()
        registers = SIMULATOR.snapshot()
        decoded = [decode_event_definition(defn, registers) for defn in definitions]

        grouped: Dict[str, List[Dict[str, Any]]] = {}
        for block in decoded:
            grouped.setdefault(block["sequence"], []).append(block)

        sequences = []
        for sequence_name, blocks in sorted(grouped.items(), key=lambda item: item[0].lower()):
            blocks.sort(key=lambda block: block["start"])
            sequences.append({"name": sequence_name, "blocks": blocks})

        return self._send_json(
            HTTPStatus.OK,
            {
                "sequences": sequences,
                "totalSequences": len(sequences),
                "totalBlocks": len(decoded),
                "notes": [
                    "Sequences come from column K of Registers_RevM_EventTypeList_WithAlarmText.",
                    "Each EventTrigger block is decoded from one C..L row (10 DM words).",
                    "eventTypeValue=first word, outputs=words[1..4], inputs=words[5..8].",
                    "Load simulator registers from PlcDeviceValue first for accurate values.",
                ],
            },
        )

    def _read_json_body(self) -> Optional[Dict[str, Any]]:
        content_len = self.headers.get("Content-Length")
        if not content_len:
            self._send_json(HTTPStatus.BAD_REQUEST, {"error": "Missing Content-Length"})
            return None
        try:
            body = self.rfile.read(int(content_len))
            payload = json.loads(body.decode("utf-8"))
            if not isinstance(payload, dict):
                raise ValueError("Expected JSON object")
            return payload
        except Exception as exc:
            self._send_json(
                HTTPStatus.BAD_REQUEST, {"error": f"Invalid JSON body: {exc}"}
            )
            return None

    def _send_json(self, status: HTTPStatus, payload: Dict[str, Any]) -> None:
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)


def run_server(port: int) -> None:
    load_default_event_definitions()
    server = ThreadingHTTPServer(("0.0.0.0", port), RequestHandler)
    print(f"PLC UI available at http://127.0.0.1:{port}")
    server.serve_forever()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Keyence PLC register UI server")
    parser.add_argument("--port", type=int, default=8080, help="HTTP server port")
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()
    run_server(args.port)
